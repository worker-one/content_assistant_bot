import logging

import openpyxl
from openpyxl.styles import Alignment

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def format_excel_file(filepath: str) -> str:
    """ Apply formatting to an Excel file

    Args:
        filepath: Path to the Excel file

    Returns:
        filepath: Path to the formatted Excel file

    """
    # Load the Excel file to apply formatting
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Set the alignment to center for all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=ws.max_column-1):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception as e:
                logger.error(f"Error formatting Excel file: {e}")
        adjusted_width = (max_length + 2)  # Add extra padding
        ws.column_dimensions[column].width = adjusted_width

    # Save the formatted Excel file
    wb.save(filepath)
    return filepath
